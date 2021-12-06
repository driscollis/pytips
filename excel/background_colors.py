# background_colors.py

from openpyxl import Workbook
from openpyxl.styles import PatternFill


def background_colors(path):
    workbook = Workbook()
    sheet = workbook.active
    yellow = "00FFFF00"
    for rows in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=12):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(start_color=yellow, end_color=yellow,
                                        fill_type = "solid")
    workbook.save(path)


if __name__ == "__main__":
    background_colors("bg.xlsx")